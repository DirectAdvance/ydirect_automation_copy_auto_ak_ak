#!/usr/bin/env python3
"""Build a deterministic gen_ses source manifest from Yandex Direct XLSX exports.

The XLSX files are treated as read-only evidence.  Campaign-level settings that are
only visible in the accompanying screenshots are recorded in ``SCREEN_META``.
Audience/retargeting condition IDs are deliberately marked as missing: the export
does not contain them and they must never be invented.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


SCREEN_META = {
    "Интересы": {
        "slug": "rsya_interests", "tp": "tp1", "engine_type": "tp1_rsy",
        "placement": "network", "weekly_budget": 35000, "status": "stopped",
        "source_ready": False, "missing": ["interest_ids_by_group"],
        "bid_modifiers": ["age_0_17=-100", "age_55_plus=-100", "tablet=-100",
                          "connected_tv=-100", "audience:Местоположение=-100"],
    },
    "Ретаргетинг+": {
        "slug": "rsya_retargeting", "tp": "tp1", "engine_type": "tp1_rsy",
        "placement": "network", "weekly_budget": 210000, "status": "active",
        "source_ready": False, "missing": ["retargeting_condition_ids_by_group"],
        "bid_modifiers": ["age_0_17=-100", "tablet=-100", "connected_tv=-100",
                          "audience:Отказы=-100", "audience:Местоположение=-100"],
    },
    "Автотаргет": {
        "slug": "rsya_autotarget", "tp": "tp1", "engine_type": "tp1_rsy",
        "placement": "network", "weekly_budget": 35000, "status": "stopped",
        "source_ready": True, "missing": [],
        "bid_modifiers": ["age_0_17=-100", "age_55_plus=-100", "tablet=-100",
                          "connected_tv=-100", "audience:Местоположение=-100"],
    },
    "Поиск": {
        "slug": "search_autotarget", "tp": "tp2", "engine_type": "search_test",
        "placement": "search+dynamic", "weekly_budget": 70000, "status": "active",
        "source_ready": True, "missing": [],
        "bid_modifiers": ["age_0_17=-100", "tablet=-100", "connected_tv=-100",
                          "audience:Местоположение=-100", "network=+25", "exclusive=+25"],
    },
    "Поиск ключи": {
        "slug": "search_keywords", "tp": "tp2", "engine_type": "search_test",
        "placement": "search+dynamic", "weekly_budget": 70000, "status": "active",
        "source_ready": True, "missing": [],
        "bid_modifiers": ["age_0_17=-100", "tablet=-100", "connected_tv=-100",
                          "audience:Местоположение=-100", "network=+25", "exclusive=+25"],
    },
    "Товарная": {
        "slug": "product_gallery", "tp": "tp3", "engine_type": "rsya_gallery",
        "placement": "product_gallery", "weekly_budget": 350000, "status": "active",
        "source_ready": True, "missing": [],
        "bid_modifiers": ["age_0_17=-100", "tablet=-100", "connected_tv=-100",
                          "audience:Местоположение=-100"],
    },
}

# Explicit spelling aliases between Direct exports and the existing coder dictionary.
# Keep this small and reviewable; fuzzy matching could silently attach the wrong ct.
GROUP_ALIASES = {
    "москвич": "moskvich",
}


def _norm(value: str) -> str:
    return re.sub(r"[^a-zа-я0-9]+", "", str(value or "").lower().replace("ё", "е"))


def _campaign_name(path: Path) -> str:
    stem = re.sub(r"\s+\d{9}(?:\s+\(остановлена\))?$", "", path.stem).strip()
    if stem not in SCREEN_META:
        raise ValueError(f"unknown gen_ses campaign export: {path.name}")
    return stem


def _campaign_id(path: Path) -> int:
    match = re.search(r"\b(\d{9})\b", path.name)
    if not match:
        raise ValueError(f"campaign id missing in {path.name}")
    return int(match.group(1))


def _current_group_map(structure_path: Path) -> dict[str, dict]:
    struct = json.loads(structure_path.read_text(encoding="utf-8"))
    directologist = next(d for d in struct["directologists"] if d.get("key") == "gen_ses")
    site = next(s for s in directologist["site_types"] if s.get("name") == "С пробегом")
    out: dict[str, dict] = {}
    for tp in site.get("tp", []):
        for group in tp.get("groups", []):
            for item in group.get("items", []):
                label = re.sub(r"^(?:РСЯ|Поиск)\s+", "", item.get("t") or "", flags=re.I)
                label = re.sub(r"\s+марка$", "", label, flags=re.I).strip()
                match = re.search(r"ct\d{4}", item.get("gc") or "")
                out.setdefault(_norm(label), {
                    "canonical_name": label,
                    "ct": match.group(0) if match else "ct0000",
                    "gc": item.get("gc") or "",
                })
    return out


def _split(value) -> list[str]:
    return [x.strip() for x in str(value or "").split("||") if x.strip()]


def read_campaign(path: Path, group_map: dict[str, dict]) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Тексты"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=14, max_row=14))]
    idx = {str(name): pos for pos, name in enumerate(headers) if name}
    groups: OrderedDict[str, dict] = OrderedDict()
    for row in ws.iter_rows(min_row=16, values_only=True):
        raw_name = row[idx["Название группы"]]
        if raw_name in (None, "", "Название группы"):
            continue
        name = str(raw_name).strip()
        lookup = GROUP_ALIASES.get(_norm(name), _norm(name))
        mapped = group_map.get(lookup, {})
        group = groups.setdefault(name, {
            "name": name,
            "canonical_name": mapped.get("canonical_name") or name,
            "ct": mapped.get("ct") or "ct0000",
            "gc": mapped.get("gc") or "",
            "autotarget": False,
            "keywords": [],
            "minus": [],
            "ads": [],
        })
        phrase = str(row[idx["Фраза (с минус-словами)"]] or "").strip()
        if phrase == "---autotargeting":
            group["autotarget"] = True
        elif phrase and phrase not in group["keywords"]:
            group["keywords"].append(phrase)
        minus = _split(row[idx["Минус-фразы на группу"]])
        for item in minus:
            if item not in group["minus"]:
                group["minus"].append(item)
        ad_id = row[idx["ID объявления"]]
        if ad_id not in (None, ""):
            ad = {
                "id": int(ad_id),
                "title": str(row[idx["Заголовок 1"]] or "").strip(),
                "title2": str(row[idx["Заголовок 2"]] or "").strip(),
                "text": str(row[idx["Текст"]] or "").strip(),
                "href": str(row[idx["Ссылка"]] or "").strip(),
                "sitelink_titles": _split(row[idx["Заголовки быстрых ссылок"]]),
                "sitelink_hrefs": _split(row[idx["Адреса быстрых ссылок"]]),
                "callouts": _split(row[idx["Уточнения"]]),
            }
            if ad not in group["ads"]:
                group["ads"].append(ad)

    name = _campaign_name(path)
    meta = dict(SCREEN_META[name])
    meta.update({
        "source_name": name,
        "source_file": path.name,
        "source_sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        "source_campaign_id": _campaign_id(path),
        "strategy": "maximum_conversions",
        "pay": "clicks",
        "counter_id": 109285853,
        "goals": [
            {"id": 560841913, "name": "Ecommerce: добавление в корзину", "value": 10},
            {"id": 560955349, "name": "Все формы", "value": 2000},
        ],
        # Project invariants intentionally override the legacy source toggles.
        "source_personalization": True,
        "source_site_monitoring": False,
        "project_personalization": False,
        "project_site_monitoring": True,
        "groups": list(groups.values()),
        "source_totals": {
            "groups": len(groups),
            "autotarget_groups": sum(bool(group["autotarget"]) for group in groups.values()),
            "keywords": sum(len(group["keywords"]) for group in groups.values()),
            "unique_keywords": len({kw for group in groups.values() for kw in group["keywords"]}),
            "ads": sum(len(group["ads"]) for group in groups.values()),
        },
    })
    return meta


def build_manifest(archive_dir: Path, structure_path: Path) -> dict:
    group_map = _current_group_map(structure_path)
    campaigns = [read_campaign(path, group_map) for path in sorted(archive_dir.glob("*.xlsx"))]
    campaigns.sort(key=lambda item: list(SCREEN_META).index(item["source_name"]))
    return {
        "schema": 1,
        "slepok": "gen_ses",
        "site_type": "С пробегом",
        # Do not persist a workstation-specific absolute path in a deployable artifact.
        "source": archive_dir.name,
        "notes": [
            "XLSX and screenshots describe six separate campaigns.",
            "Interest and retargeting condition IDs are absent; source_ready=false until read from Grid/API.",
            "Legacy personalization/site-monitoring values are recorded but project invariants win.",
        ],
        "campaigns": campaigns,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("archive_dir", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--structure", type=Path,
                        default=Path(__file__).resolve().parents[1] / "slepki_structure.json")
    args = parser.parse_args()
    manifest = build_manifest(args.archive_dir, args.structure)
    args.output.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
